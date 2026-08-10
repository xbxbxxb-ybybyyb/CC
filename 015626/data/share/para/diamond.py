import os
import shutil
import pandas as pd
from openpyxl import load_workbook

source_date = 20250612
target_date = 20250613

# 源路径和目标路径
source_dir = f"/data/user/011477/Trade_Docs/{source_date}/Diamond_{source_date}/"  # 替换为实际路径
target_dir = f"/data/user/011477/Trade_Docs/{target_date}/Diamond_{target_date}/"  # 替换为实际路径

# 创建目标路径（如果不存在）
os.makedirs(target_dir, exist_ok=True)

for filename in os.listdir(source_dir):
    if filename.endswith(".xlsx"):
        src_path = os.path.join(source_dir, filename)
        dst_path = os.path.join(target_dir, filename)

        # 复制文件
        shutil.copy(src_path, dst_path)

        try:
            # 加载 Excel 文件
            wb = load_workbook(dst_path)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            # 查找“交易日期”列的位置
            headers = [cell.value for cell in ws[1]]
            if "交易日" in headers:
                col_idx = headers.index("交易日") + 1  # 列索引从1开始

                # 修改“交易日期”列的数据（从第二行开始）
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    cell.value = target_date

                # 保存修改
                wb.save(dst_path)
                print(f"✅ 文件 {filename} 的第一个 Sheet 已成功修改。")
            else:
                print(f"⚠️ 文件 {filename} 中没有 '交易日' 列，跳过处理。")

        except Exception as e:
            print(f"❌ 处理文件 {filename} 时出错：{e}")
import datetime as dt
import pandas as pd
import os
import numpy as np
import json
from WindPy import w

w.start()
pd.set_option('max_columns', 30)

from openpyxl import load_workbook

def excelAddSheet(dataframe, excelWriter, sheet_name):
    book = load_workbook(excelWriter.path)
    excelWriter.book = book

    sheetlist = [x.title for x in book.worksheets]
    if sheet_name in sheetlist:
        print(sheet_name, 'True')
        idx = book.sheetnames.index(sheet_name)
        book.remove(book.worksheets[idx])

    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheet_name)
    excelWriter.close()

rootpath = 'A:/data/share/LOCAL_DATA/COMMODITY/impacts_data/oil_precious_metal/'
filepath = rootpath + 'oil_precious_metal.xlsx'
writer = pd.ExcelWriter(filepath,engine='openpyxl')
starttime, endtime = "2015-01-01", "2030-03-30"

writer = pd.ExcelWriter(rootpath + 'oil_precious_metal.xlsx',engine='openpyxl')

#美元指数
df = w.edb("M0000271,S0031550,S0031553", starttime, endtime,"Fill=Previous")
df = pd.DataFrame(data = df.Data, index=df.Codes, columns=df.Times).T
df.columns = ['dollar_index','BDI','BDTI']
df = df.sort_index(ascending=False)
excelAddSheet(df, writer, '指数_日频')

df = w.edb("G0000035,G0000027", starttime, endtime,"Fill=Previous")
df = pd.DataFrame(data = df.Data, index=df.Codes, columns=df.Times).T
df.columns = ["CPI","CPI_yoy(%)"]
df = df.sort_index(ascending=False)
excelAddSheet(df, writer, '美国CPI及同比_月频')

# 美国当周初次申请失业金人数 人 周
df = w.edb("G0002433", starttime, endtime,"Fill=Previous")
df = pd.DataFrame(data = df.Data, index=df.Codes, columns=df.Times).T
df.columns = ["num"]
df = df.sort_index(ascending=False)
excelAddSheet(df, writer, '美国当周初次申请失业金人数_人_周频')


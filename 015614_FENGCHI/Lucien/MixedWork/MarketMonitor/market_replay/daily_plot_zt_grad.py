# coding: utf-8
# Author：fengchi863
# Date ：2024/7/30 16:56

import sys
sys.path.append('/data/user/015614/Lucien')

from dataApi import getData, stockList, tradeDate, indName
from xquant.factordata import FactorData
import matplotlib.pyplot as plt
plt.rcParams['font.family'] = ['sans-serif']
plt.rcParams['font.sans-serif'] = ['SimHei']
import numpy as np
import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from dataApi.sendInfo import send_message, send_file

fd = FactorData()

zt_data_path = '/data/group/800463/日内强势股/实盘分析记录/日内强势股成交记录/连板天梯图/'

def forward_fill(arr, axis, zero_fill=True):
    arr = arr.swapaxes(axis, -1)
    if zero_fill:
        mask = arr == 0
    else:
        mask = np.isnan(arr)
    idx = np.where(~mask, np.arange(mask.shape[-1]), 0)
    np.maximum.accumulate(idx, axis=-1, out=idx)

    out = arr[tuple(np.arange(idx.shape[x])[(None,) * x + (slice(None),) + (None,) * (idx.ndim - x - 1)] for x in range(idx.ndim - 1)) + (idx,)]
    out = out.swapaxes(axis, -1)
    return out

def get_lb(zt_flag):
    zt_values_copy = zt_flag.values.copy()
    zt_values2 = zt_values_copy.cumsum(axis=1)
    breaks = zt_values2 * (zt_values_copy == 0)
    zt_values3 = forward_fill(breaks, axis=1)
    zt_values4 = zt_values2 - zt_values3
    return zt_values4

if __name__ == '__main__':
    today_date = dt.datetime.now().strftime('%Y%m%d')
    # today_date = '20250506'
    start_date = fd.tradingday(today_date, -6)[0]
    start_date, end_date = int(start_date), int(today_date)
    week_start_date, week_end_date = start_date, end_date
    Friday = week_end_date
    Friday_str = pd.to_datetime(str(Friday)).strftime('%Y-%m-%d')

    shift_start_date = tradeDate.get_pre_trade_date(start_date, 40)
    date_list = tradeDate.get_date_range(start_date, end_date)
    shift_date_list = tradeDate.get_date_range(shift_start_date, end_date)
    week_date_list = tradeDate.get_date_range(week_start_date, week_end_date)

    jupiter_data = pd.DataFrame()
    for week_date in week_date_list:
        tmp_jupiter_data = pd.read_pickle(f'/data/user/015614/daily/basic/basic_wind_sw_history3/BlockData/daily_max_pctchg_concept/jupiter/{week_date}.pkl')
        jupiter_data = pd.concat([jupiter_data, tmp_jupiter_data], axis=0)
    jupiter_dict = jupiter_data['概念名称'].reset_index().set_index('Ticker')['概念名称'].to_dict()

    stk_pool = stockList.clean_stock_list(no_ST=True, least_live_days=10, least_normal_days=10, no_pause=True, least_recover_days=0,
                                          start_date=shift_start_date, end_date=end_date)
    stk_list = stk_pool.iloc[-1].index.tolist()

    sw2 = getData.get_daily_1factor('SW20212', code_list=stk_list, date_list=[Friday]).T.reset_index().set_index('code')[Friday]
    sw2 = sw2.apply(lambda x: indName.sw2021_level2[x])
    sw2_dict = sw2.to_dict()

    def format_stock(stk_id, stk_name, jupiter_dict=jupiter_dict, sw2_dict=sw2_dict):
        stk_code = stockList.trans_int2windcode(stk_id)
        return f'{stk_name}({jupiter_dict[stk_code].split(",")[0] if (stk_code in jupiter_dict.keys()) and (type(jupiter_dict[stk_code]) == str) else ""})'

    limit_max = getData.get_daily_1factor('limit_max', date_list=shift_date_list, code_list=stk_list) # 属于不复权的价格
    limit_min = getData.get_daily_1factor('limit_min', date_list=shift_date_list, code_list=stk_list)
    daily_pctchg = getData.get_daily_1factor('pct_chg', date_list=shift_date_list, code_list=stk_list)
    opn = getData.get_daily_1factor('open', date_list=shift_date_list, code_list=stk_list)
    low = getData.get_daily_1factor('low', date_list=shift_date_list, code_list=stk_list)
    high = getData.get_daily_1factor('high', date_list=shift_date_list, code_list=stk_list)
    close = getData.get_daily_1factor('close', date_list=shift_date_list, code_list=stk_list)
    high_badj = getData.get_daily_1factor('high_badj', date_list=shift_date_list, code_list=stk_list)
    open_badj = getData.get_daily_1factor('open_badj', date_list=shift_date_list, code_list=stk_list)
    pre_close_badj = getData.get_daily_1factor('pre_close_badj', date_list=shift_date_list, code_list=stk_list)
    daily_high_pctchg = (high_badj / pre_close_badj - 1) * 100
    daily_open_pctchg = (open_badj / pre_close_badj - 1) * 100
    zt = pd.DataFrame((close == limit_max)) & stk_pool
    filter_zt = pd.DataFrame((close == limit_max) & (opn != limit_max) & (close.shift(1) != limit_max.shift(1)) & stk_pool) # 去掉一字板和T字板
    dt = pd.DataFrame(close == limit_min) & stk_pool
    zb = pd.DataFrame(((close != limit_max) & (high == limit_max))) & stk_pool
    filter_zb = pd.DataFrame((close != limit_max) & (high == limit_max) & (opn != limit_max)) & stk_pool

    zt = zt & (daily_pctchg > 6)
    filter_zt = filter_zt & (daily_pctchg > 6)
    dt = dt & (daily_pctchg < -6)
    zb = zb & (daily_high_pctchg > 6)   # 只要触过板就算，但同花顺上比如20220812宇通重工、宏和科技就不算在内，应该是必须上板一段时间
    filter_zb = filter_zb & (daily_high_pctchg > 6)

    short_name = fd.get_factor_value('Basic_factor', mddate=[str(end_date)], factor_names=['short_name'])
    short_name.index = short_name.index.map(stockList.trans_windcode2int)
    short_name = short_name.to_dict()['short_name']

    lb = pd.DataFrame(get_lb(zt.T).T, index=zt.index, columns=zt.columns)
    lb_height = lb.max(axis=1)  # 连板高度

    lb = lb.loc[date_list]
    lb_height = lb_height[date_list]

    # 计算连板梯度
    stats_dict = {}
    lb_grad = pd.DataFrame(index=range(2, 15)[::-1], columns=week_date_list)
    for _date in week_date_list:
        for _grad in lb_grad.index:
            if _grad in lb.T[_date].tolist():
                lb_list = lb.loc[_date][lb.loc[_date] == _grad].index.tolist()
                # 这里根据排序指标构造Series，然后排序后取index
                ind_series = pd.Series(dict(zip(lb_list, list(map(lambda x: sw2[x] if x in sw2.keys() else np.nan, lb_list))))).sort_values()
                lb_list = ind_series.index.tolist()
                format_lb_list = '\n'.join(list(map(lambda x: format_stock(x, short_name[x]), lb_list)))
                lb_grad.loc[_grad, _date] = format_lb_list
    _lb_grad = lb_grad.loc[:, week_start_date:week_end_date]
    _lb_grad = _lb_grad.dropna(how='all', axis=0)
    _lb_grad.columns = _lb_grad.columns.astype(str)
    _lb_grad.index = _lb_grad.index.astype(str)
    # _lb_grad.to_excel(zt_data_path + f'{today_date}.xlsx')

    writer = pd.ExcelWriter(zt_data_path + f'{today_date}.xlsx', engine='xlsxwriter')
    _lb_grad.to_excel(writer, sheet_name='Sheet1', index=True)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    worksheet.set_column('A:G', 25)
    writer.close()

    wb = load_workbook(zt_data_path + f'{today_date}.xlsx')
    ws = wb.active

    for cells in ws.rows:
        for cell in cells:
            cell.alignment = Alignment(wrap_text=True)
            if cell.column == 'A':
                cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(zt_data_path + f'{today_date}.xlsx')

    # %% 这种绘图方式不好调整行高等
    # # 读取 Excel 文件
    # data = pd.read_excel(zt_data_path + f'20240730.xlsx')
    # data = data.fillna('')
    # plt.figure(figsize=(20, 40))
    # plt.table(cellText=data.values, colLabels=data.columns, loc='center')
    # plt.axis('off')
    #
    # plt.savefig(zt_data_path + 'test.png')

    # %% 第二种绘图方式
    from PIL import Image, ImageFont, ImageDraw

    wb = load_workbook(f'{zt_data_path}{today_date}.xlsx')
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df = df.fillna(' ')

    adapt_perrow_width = df.iloc[-2].apply(lambda x: len(x.split('\n'))).max()
    row_width = 20 * adapt_perrow_width  # 行高
    column_width = 140  # 列宽

    img = Image.new('RGB', (column_width * sheet.max_column, row_width * sheet.max_row + 200), color='white')  # 肯定是7列，除非加1天
    font = ImageFont.truetype("/data/user/015614/python_package/simhei.ttf", 12)
    big_font = ImageFont.truetype("/data/user/015614/python_package/simhei.ttf", 24)
    draw = ImageDraw.Draw(img)
    colors_list = [(0, 0, 0), (162, 189, 244)]

    draw.rectangle([(0, 0), (column_width - 1, row_width * sheet.max_row + 200)], fill=colors_list[1])
    draw.rectangle([(0, 0), (column_width * sheet.max_column, row_width - 1)], fill=colors_list[1])

    for y in range(len(df)):
        for x in range(len(df.columns)):
            if x != 0 and y != 0:
                draw.text((column_width * x, row_width * y), str(df.iloc[y, x]), font=font, fill=colors_list[0])
            else:
                draw.text((column_width * x, row_width * y), str(df.iloc[y, x]), font=big_font, fill=colors_list[0])

    for i in range(1, sheet.max_row):
        draw.line((0, i * row_width - 1, sheet.max_column * column_width, i * row_width - 1), fill='gray')
    for i in range(1, sheet.max_column):
        draw.line((i * column_width - 1, 0, i * column_width - 1, sheet.max_row * row_width + 200), fill='gray')

    img.save(f'{zt_data_path}{today_date}.png', )

    send_message(f'{today_date}连板天梯图已保存完毕')
    send_file(f'{zt_data_path}{today_date}.png')